from openpyxl import load_workbook
import pandas as pd
from copy import deepcopy
import ast

def appendToExcel(outputfile, sheetName, df1):
    try:
        book = load_workbook(outputfile)
        writer = pd.ExcelWriter(outputfile, engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        df1.to_excel(writer,sheet_name=sheetName, startrow=writer.sheets[sheetName].max_row, index = False,header= False)
        writer.save()
    except FileNotFoundError:
        with pd.ExcelWriter(outputfile) as writer: 
            df1.to_excel(writer, sheet_name=sheetName, header = True, index = False)

def writeToExcel(outputfile, sheetName, df1):
    try:
        book = load_workbook(outputfile)
        writer = pd.ExcelWriter(outputfile, engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        df1.to_excel(writer, sheet_name=sheetName, index = False, header= False)
        writer.save()
    except FileNotFoundError:
        with pd.ExcelWriter(outputfile) as writer: 
            df1.to_excel(writer, sheet_name=sheetName, header = False, index = False)
 

def eval_df(df):
    res = deepcopy(df)
    for i in range(df.shape[0]):
        for j in range(df.shape[1]):
            res.iloc[i,j] = ast.literal_eval(res.iloc[i,j])
    return res

